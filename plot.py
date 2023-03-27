import matplotlib.pyplot as plt
import openpyxl

workbook_dropout1 = openpyxl.load_workbook('./excel/dropout_1.xlsx')
workbook_dropout2 = openpyxl.load_workbook('./excel/dropout_2.xlsx')
workbook_dense = openpyxl.load_workbook('./excel/dense.xlsx')
workbook_conv1 = openpyxl.load_workbook('./excel/conv_1.xlsx')
workbook_conv2 = openpyxl.load_workbook('./excel/conv_2.xlsx')
sheet_dropout1 = workbook_dropout1.active
sheet_dropout2 = workbook_dropout2.active
sheet_dense = workbook_dense.active
sheet_conv1 = workbook_conv1.active
sheet_conv2 = workbook_conv2.active

acc_dropout_1 = []
acc_dropout_2 = []
acc_dense = []
acc_conv_1 = []
acc_conv_2 = []
epoch_acc_dropout_1 = []
epoch_acc_dropout_2 = []
epoch_acc_dense = []
epoch_acc_conv_1 = []
epoch_acc_conv_2 = []


for i in range(7):
    acc_dropout_1.append(sheet_dropout1.cell(row=2, column=i+2).value)
    acc_dropout_2.append(sheet_dropout2.cell(row=2, column=i+2).value)
    acc_dense.append(sheet_dense.cell(row=2, column=i+2).value)
    acc_conv_1.append(sheet_conv1.cell(row=2, column=i+2).value)
    acc_conv_2.append(sheet_conv2.cell(row=2, column=i+2).value)
    epoch_acc_dropout_2.append(sheet_dropout2.cell(row=3, column=i+2).value)
    epoch_acc_dropout_2.append(sheet_dropout2.cell(row=3, column=i+2).value)
    epoch_acc_dense.append(sheet_dense.cell(row=3, column=i+2).value)
    epoch_acc_conv_1.append(sheet_conv1.cell(row=3, column=i+2).value)
    epoch_acc_conv_2.append(sheet_conv2.cell(row=3, column=i+2).value)
    
tested_values_dropout = [0, 0.05, 0.1, 0.15, 0.2, 0.25, 0.3]
tested_values_dense_conv = [8, 16, 32, 64, 128, 256, 512]

plt.plot(tested_values_dense_conv, acc_conv_2, label='Accuracy')
#plt.plot(tested_values, epoch_acc_dropout_1, label='Epoch Accuracy')
plt.xscale('log', base=2)
plt.xticks(tested_values_dense_conv)
plt.xlabel('Size of the convolutional layer')
plt.ylabel('Accuracy')
plt.title('Accuracy for the convolutional layer C2')
plt.legend()
plt.savefig('images/conv_2.png')
plt.show()