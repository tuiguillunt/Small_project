import openpyxl
import subprocess
import os
import string

file_path = './excel/conv_1.xlsx'
alphabet = list(string.ascii_uppercase)

val_to_determine = 'conv_1'
tested_values = [8, 16, 32, 64, 128, 256, 512]
current_tested_value = 6

column_letter = alphabet[current_tested_value + 1]
string_acc = ''.join([column_letter, str(2)])
string_epoch_acc = ''.join([column_letter, str(3)])

def initialize_excel(sheet):
    sheet['A1'] = val_to_determine
    for i in range(len(tested_values)):
        sheet.cell(row=1, column=i+2, value=tested_values[i])
    sheet['A2'] = 'accuracy'
    sheet['A3'] = 'last epoch accuracy'

if os.path.exists(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    if sheet.cell(row=3, column=len(tested_values) + 1).value is not None:
        os.remove(file_path)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        initialize_excel(sheet)
else:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    initialize_excel(sheet)
# List
accuracy_list = []
epoch_accuracy_list = []
for_range = 20
for i in range(for_range):
    print(i, '/', for_range)
    result = subprocess.run(['python', 'projet.py'], capture_output=True, text=True)
    test_acc = float(result.stdout.strip().split()[-2])
    test_epoch_acc = float(result.stdout.strip().split()[-1])
    accuracy_list.append(test_acc)
    epoch_accuracy_list.append(test_epoch_acc)

mean_acc = sum(accuracy_list) / len(accuracy_list)
print('mean acc: ', mean_acc)
mean_epoch_acc = sum(epoch_accuracy_list) / len(epoch_accuracy_list)
print('mean epoch acc: ', mean_epoch_acc)

sheet[string_acc] = mean_acc
sheet[string_epoch_acc] = mean_epoch_acc

# Enregistrer le classeur Excel
excel_name = ''.join(['excel/', val_to_determine, '.xlsx'])
workbook.save(excel_name)